"""
Excel/CSV converter for AI-Ready File Converter.
Extracts tabular data from spreadsheets, optimized for JSON output.
"""

from pathlib import Path
import csv
from openpyxl import load_workbook
from .base import BaseConverter


class XlsxConverter(BaseConverter):
    """Converter for Excel and CSV files."""
    
    def _get_file_type(self) -> str:
        ext = self.file_path.suffix.lower()
        if ext == '.csv':
            return "csv_spreadsheet"
        return "excel_spreadsheet"
    
    def _extract_content(self) -> dict:
        """Extract content from spreadsheet."""
        ext = self.file_path.suffix.lower()
        
        if ext == '.csv':
            return self._extract_csv()
        else:
            return self._extract_xlsx()
    
    def _extract_csv(self) -> dict:
        """Extract content from CSV file."""
        content = {
            "sheets": [],
            "text": ""
        }
        
        rows = []
        headers = []
        
        with open(self.file_path, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i == 0:
                    headers = row
                rows.append(row)
        
        sheet_data = {
            "name": "Sheet1",
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
            "column_count": len(headers) if headers else 0
        }
        
        content["sheets"].append(sheet_data)
        
        # Build text representation
        text_parts = []
        for row in rows:
            text_parts.append(", ".join(str(cell) for cell in row))
        content["text"] = "\n".join(text_parts)
        
        # Update metadata
        self._metadata = {
            "sheet_count": 1,
            "total_rows": len(rows),
            "total_columns": len(headers) if headers else 0
        }
        
        return content
    
    def _extract_xlsx(self) -> dict:
        """Extract content from Excel file."""
        content = {
            "sheets": [],
            "text": ""
        }
        
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        
        total_rows = 0
        text_parts = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            rows = []
            headers = []
            
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                # Convert cells to strings, handling None values
                row_data = [str(cell) if cell is not None else "" for cell in row]
                
                # Skip completely empty rows
                if not any(row_data):
                    continue
                
                if i == 0:
                    headers = row_data
                rows.append(row_data)
                text_parts.append(", ".join(row_data))
            
            sheet_data = {
                "name": sheet_name,
                "headers": headers,
                "rows": rows,
                "row_count": len(rows),
                "column_count": len(headers) if headers else 0
            }
            
            content["sheets"].append(sheet_data)
            total_rows += len(rows)
        
        wb.close()
        
        content["text"] = "\n".join(text_parts)
        
        # Update metadata
        self._metadata = {
            "sheet_count": len(wb.sheetnames),
            "total_rows": total_rows
        }
        
        return content
    
    def _format_as_markdown(self, content: dict) -> str:
        """Format spreadsheet content as Markdown tables."""
        md_parts = []
        
        for sheet in content["sheets"]:
            # Sheet header
            md_parts.append(f"## {sheet['name']}")
            md_parts.append("")
            
            if not sheet["rows"]:
                md_parts.append("*Empty sheet*")
                md_parts.append("")
                continue
            
            # Determine column widths for better formatting
            if sheet["headers"]:
                # Header row
                md_parts.append("| " + " | ".join(sheet["headers"]) + " |")
                md_parts.append("| " + " | ".join(["---"] * len(sheet["headers"])) + " |")
                
                # Data rows (skip first row as it's headers)
                for row in sheet["rows"][1:]:
                    # Ensure row has same number of columns as headers
                    padded_row = row + [""] * (len(sheet["headers"]) - len(row))
                    md_parts.append("| " + " | ".join(padded_row[:len(sheet["headers"])]) + " |")
            else:
                # No headers - just output rows
                for row in sheet["rows"]:
                    md_parts.append("| " + " | ".join(row) + " |")
            
            md_parts.append("")
            md_parts.append("---")
            md_parts.append("")
        
        return "\n".join(md_parts)
    
    def _build_json_structure(self, content: dict) -> dict:
        """Build JSON structure optimized for structured data."""
        # For spreadsheets, we provide a more data-focused structure
        from datetime import datetime, timezone
        
        sheets_data = []
        for sheet in content["sheets"]:
            # Convert rows to objects using headers as keys
            if sheet["headers"] and len(sheet["rows"]) > 1:
                records = []
                for row in sheet["rows"][1:]:  # Skip header row
                    record = {}
                    for i, header in enumerate(sheet["headers"]):
                        if header:  # Only include non-empty headers
                            record[header] = row[i] if i < len(row) else ""
                    records.append(record)
                
                sheets_data.append({
                    "name": sheet["name"],
                    "headers": sheet["headers"],
                    "records": records,
                    "raw_rows": sheet["rows"]
                })
            else:
                sheets_data.append({
                    "name": sheet["name"],
                    "raw_rows": sheet["rows"]
                })
        
        return {
            "source": {
                "filename": self.filename,
                "type": self.file_type,
                "converted_at": datetime.now(timezone.utc).isoformat()
            },
            "content": {
                "sheets": sheets_data
            },
            "metadata": self._metadata
        }
